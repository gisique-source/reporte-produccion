"""Exportación mensual a Excel (formato planta + gráficos en hoja Resumen)."""

from __future__ import annotations

import calendar
from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Side

from codigos_produccion import codigos_desde_registro
from db import PesajeDatabase, nombre_mes
from export_styles import (
    BORDER,
    FILL_KPI,
    FILL_KPI2,
    FILL_KPI3,
    FILL_KPI4,
    FILL_TOTAL,
    FONT_META,
    FONT_TITLE,
    ajustar_columnas_dia,
    aplicar_fila_produccion,
    estilo_celda_header,
    estilo_fila_resumen,
    insertar_logo,
)
from models import RegistroPesaje, ResumenDia

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

_MESES_ABR = (
    "",
    "ENE",
    "FEB",
    "MAR",
    "ABR",
    "MAY",
    "JUN",
    "JUL",
    "AGO",
    "SEP",
    "OCT",
    "NOV",
    "DIC",
)

_HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
_LIGHT_FILL = PatternFill("solid", fgColor="DDEBF7")
_TOTAL_FILL = FILL_TOTAL
_THIN = Side(style="thin", color="B4C6E7")

_DIA_HEADERS: tuple[tuple[int, str], ...] = (
    (2, "Fardo"),
    (3, "Cliente"),
    (4, "Lote"),
    (5, "Color"),
    (6, "Dn:"),
    (7, "Corte mm"),
    (8, "P. Total"),
    (9, "Tara Carreta"),
    (10, "Tara Fardo"),
    (11, "P. Bruto"),
    (12, "P. Neto"),
    (13, "Hora"),
    (14, "Operario"),
    (15, "Beteado"),
    (16, "Verificacion"),
)


def nombre_archivo_mensual(year: int, month: int, ext: str = "xlsx") -> str:
    """Nombre tipo planta: ETIQUETA EXTRUSORA 01-26- ENE.xlsx"""
    abr = _MESES_ABR[month] if 1 <= month <= 12 else f"{month:02d}"
    return f"ETIQUETA EXTRUSORA {month:02d}-{year % 100:02d}- {abr}.{ext}"


def _hora(reg: RegistroPesaje) -> str:
    try:
        dt = datetime.strptime(reg.fecha_hora, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%H:%M")
    except ValueError:
        return reg.fecha_hora[11:16] if len(reg.fecha_hora) >= 16 else ""


def _style_header_row(ws: "Worksheet", row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        estilo_celda_header(ws.cell(row=row, column=col))


def _escribir_resumen(
    ws: "Worksheet",
    *,
    year: int,
    month: int,
    filas: Sequence[ResumenDia],
) -> None:
    ws.title = "Resumen"
    days = len(filas)
    tb = sum(r.peso_bruto for r in filas)
    tn = sum(r.peso_neto for r in filas)
    tf = sum(r.cantidad for r in filas)
    td = sum(1 for r in filas if r.cantidad > 0)

    insertar_logo(ws, "A1")
    ws.merge_cells("C2:H2")
    ws["C2"] = "GEXIM S.A.C. · PRECIX-WEIGHT"
    ws["C2"].font = Font(bold=True, size=12, color="1F4E79")

    ws.merge_cells("C4:H4")
    ws["C4"] = "Resumen de Producción Mensual — Sección Extrusora"
    ws["C4"].font = Font(bold=True, size=14, color="C00000")

    ws.merge_cells("C5:H5")
    ws["C5"] = f"Periodo: {nombre_mes(month)} {year}"
    ws["C5"].font = Font(size=11, color="44546A")

    kpi_fills = (FILL_KPI, FILL_KPI2, FILL_KPI3, FILL_KPI4)
    kpis = (
        ("Fardos", tf),
        ("Peso bruto (kg)", round(tb, 1)),
        ("Peso neto (kg)", round(tn, 1)),
        ("Días con producción", td),
    )
    for i, (lbl, val) in enumerate(kpis):
        col = 2 + i * 2
        fill = kpi_fills[i]
        ws.cell(7, col, lbl).font = Font(bold=True, size=9, color="44546A")
        ws.cell(7, col).fill = fill
        ws.cell(7, col).alignment = Alignment(horizontal="center")
        ws.cell(7, col).border = BORDER
        c = ws.cell(8, col, val)
        c.font = Font(bold=True, size=14, color="1F4E79")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    hdr_row = 10
    headers = ("Día", "Fecha", "Fardos", "P. Bruto (kg)", "P. Neto (kg)")
    for col, h in enumerate(headers, start=2):
        ws.cell(hdr_row, col, h)
    _style_header_row(ws, hdr_row, len(headers) + 1)

    data_start = hdr_row + 1
    for i, r in enumerate(filas):
        row = data_start + i
        ws.cell(row, 2, r.dia).alignment = Alignment(horizontal="center")
        ws.cell(row, 3, r.fecha)
        ws.cell(row, 4, r.cantidad).alignment = Alignment(horizontal="center")
        ws.cell(row, 5, round(r.peso_bruto, 1) if r.peso_bruto else 0)
        ws.cell(row, 6, round(r.peso_neto, 1) if r.peso_neto else 0)
        if r.cantidad > 0:
            ws.cell(row, 5).font = Font(color="C00000")
            ws.cell(row, 6).font = Font(bold=True)
        estilo_fila_resumen(ws, row, produccion=r.cantidad > 0, par=i % 2 == 0)

    total_row = data_start + days
    ws.cell(total_row, 2, "TOTAL MES").font = Font(bold=True)
    ws.cell(total_row, 4, tf).font = Font(bold=True)
    ws.cell(total_row, 5, round(tb, 1)).font = Font(bold=True)
    ws.cell(total_row, 6, round(tn, 1)).font = Font(bold=True)
    for col in range(2, 7):
        ws.cell(total_row, col).fill = _TOTAL_FILL
        ws.cell(total_row, col).border = BORDER

    data_end = data_start + days - 1
    cats = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    bruto_ref = Reference(
        ws, min_col=5, min_row=hdr_row, max_row=data_end
    )
    neto_ref = Reference(
        ws, min_col=6, min_row=hdr_row, max_row=data_end
    )

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Producción diaria — P. Neto (kg)"
    bar.y_axis.title = "kg"
    bar.x_axis.title = "Día del mes"
    bar.style = 10
    bar.add_data(neto_ref, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 9
    bar.width = 16
    ws.add_chart(bar, "H4")

    line = LineChart()
    line.title = "P. Bruto vs P. Neto (kg)"
    line.y_axis.title = "kg"
    line.x_axis.title = "Día del mes"
    line.add_data(bruto_ref, titles_from_data=True)
    line.add_data(neto_ref, titles_from_data=True)
    line.set_categories(cats)
    line.height = 9
    line.width = 16
    ws.add_chart(line, "H20")

    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14


def _escribir_metadatos_dia(ws: "Worksheet", dia: date) -> None:
    ws["S1"] = dia.year
    ws["T1"] = dia.month
    ws["S3"] = dia.year % 100
    ws["T3"] = f"{dia.month:02d}"


def _escribir_codigos_fila(
    ws: "Worksheet",
    row: int,
    reg: RegistroPesaje,
    dia: date,
    catalogo,
) -> dict[str, str]:
    """Columnas auxiliares O–AC (misma lógica que el Excel de planta)."""
    cod = codigos_desde_registro(
        reg, catalogo=catalogo, anio=dia.year, mes=dia.month
    )
    ws.cell(row, 15, reg.beteado or "1")  # O beteado
    ws.cell(row, 19, cod["color_codigo"])  # S
    ws.cell(row, 20, cod["dn_3"])  # T
    ws.cell(row, 22, cod["corte_3"])  # V
    ws.cell(row, 23, cod["fardo_3"])  # W
    ws.cell(row, 26, cod["bloque_z"])  # Z
    ws.cell(row, 27, cod["principal"])  # AA
    ws.cell(row, 28, cod["neto"])  # AB
    ws.cell(row, 29, cod["largo"])  # AC
    return cod


def _escribir_hoja_dia(
    ws: "Worksheet",
    dia: date,
    regs: Sequence[RegistroPesaje],
    catalogo,
) -> None:
    ws.title = f"Dia {dia.day:02d}"
    insertar_logo(ws, "A1")
    _escribir_metadatos_dia(ws, dia)

    ws.merge_cells("C5:H5")
    ws["C5"] = "HOJA DE PRODUCCION SECCION EXTRUSORA"
    ws["C5"].font = FONT_TITLE
    ws["J5"] = "Fecha de Produccion"
    ws["J5"].font = FONT_META
    ws["M5"] = dia
    ws["M5"].number_format = "DD/MM/YYYY"

    activos = [r for r in regs if r.activo]
    if activos:
        ws["M3"] = activos[-1].nro_fardo

    for col, title in _DIA_HEADERS:
        advertencia = title.lower().startswith("verific")
        estilo_celda_header(ws.cell(7, col, title), advertencia=advertencia)
    estilo_celda_header(ws.cell(7, 27, "Cod.Principal"))
    estilo_celda_header(ws.cell(7, 29, "Cod.Largo"))

    for i, reg in enumerate(activos, start=1):
        row = 7 + i
        aplicar_fila_produccion(ws, row, reg, i, hora=_hora(reg))
        _escribir_codigos_fila(ws, row, reg, dia, catalogo)

    # Vista etiqueta del último fardo (celdas F39, H40, P64 como plantilla)
    if activos:
        ult = activos[-1]
        cod_u = codigos_desde_registro(
            ult, catalogo=catalogo, anio=dia.year, mes=dia.month
        )
        ws.cell(39, 6, cod_u["principal"]).font = Font(bold=True, size=11)
        ws.cell(40, 4, ult.nro_fardo)
        ws.cell(40, 6, cod_u["fardo_3"])
        ws.cell(40, 8, cod_u["largo"]).font = Font(size=9)
        ws.cell(43, 4, ult.color)
        ws.cell(43, 6, cod_u["color_codigo"])
        ws.cell(64, 16, cod_u["principal"]).font = Font(bold=True, size=12)

    total_row = max(38, 8 + len(activos))
    bruto_t = sum(r.peso_bruto for r in activos)
    neto_t = sum(r.peso_neto for r in activos)
    ws.cell(total_row, 10, "Total del día").font = Font(bold=True, color="1F4E79")
    ws.cell(total_row, 11, round(bruto_t, 2)).font = Font(bold=True, color="C00000")
    ws.cell(total_row, 12, round(neto_t, 2)).font = Font(bold=True)
    for col in range(10, 13):
        ws.cell(total_row, col).fill = _TOTAL_FILL
        ws.cell(total_row, col).border = BORDER

    ajustar_columnas_dia(ws)


def exportar_mensual_excel(
    db: PesajeDatabase,
    year: int,
    month: int,
    destino: str | Path,
) -> Path:
    """Genera libro mensual: hoja Resumen con gráficos + hojas Dia 01…31."""
    path = Path(destino)
    resumen = db.resumen_mes(year, month)
    days_in_month = calendar.monthrange(year, month)[1]

    wb = Workbook()
    ws_res = wb.active
    assert ws_res is not None
    _escribir_resumen(ws_res, year=year, month=month, filas=resumen)

    for d in range(1, days_in_month + 1):
        dia = date(year, month, d)
        regs = db.por_fecha(dia, incluir_ocultos=False)
        ws = wb.create_sheet(title=f"Dia {d:02d}")
        _escribir_hoja_dia(ws, dia, regs, db.catalogo)

    wb.save(path)
    return path


def resumen_exportacion(db: PesajeDatabase, year: int, month: int) -> dict:
    """KPIs para la UI de exportación."""
    filas = db.resumen_mes(year, month)
    return {
        "fardos": sum(r.cantidad for r in filas),
        "bruto": sum(r.peso_bruto for r in filas),
        "neto": sum(r.peso_neto for r in filas),
        "dias_prod": sum(1 for r in filas if r.cantidad > 0),
        "filas": filas,
    }
