"""Estilos y colores para exportación Excel (filas por color de producto, logo)."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from models import RegistroPesaje
from utils import resource_path

# --- Paleta (referencia planta Extrusora) ---------------------------------

FILL_HEADER = PatternFill("solid", fgColor="5B9BD5")
FILL_HEADER_WARN = PatternFill("solid", fgColor="FF0000")
FILL_PROD = PatternFill("solid", fgColor="FFF2CC")
FILL_PESO = PatternFill("solid", fgColor="DDEBF7")
FILL_HORA = PatternFill("solid", fgColor="DAEEF3")
FILL_TOTAL = PatternFill("solid", fgColor="BDD7EE")
FILL_KPI = PatternFill("solid", fgColor="E2EFDA")
FILL_KPI2 = PatternFill("solid", fgColor="FCE4D6")
FILL_KPI3 = PatternFill("solid", fgColor="DDEBF7")
FILL_KPI4 = PatternFill("solid", fgColor="E4DFEC")

FONT_TITLE = Font(bold=True, size=13, color="C00000", underline="single")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=9)
FONT_HEADER_WARN = Font(bold=True, color="FFFFFF", size=9)
FONT_META = Font(bold=True, size=9, color="1F4E79")

_THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Palabra clave en maestro Color → hex de fuente (ARGB sin #)
_COLOR_PALETTE: tuple[tuple[str, str], ...] = (
    ("azulino", "0070C0"),
    ("azul", "0070C0"),
    ("celeste", "00B0F0"),
    ("turquesa", "00B0B0"),
    ("verde", "00B050"),
    ("amarillo", "BF8F00"),
    ("naranja", "ED7D31"),
    ("naranjo", "ED7D31"),
    ("rojo", "C00000"),
    ("marron", "843C0C"),
    ("marrón", "843C0C"),
    ("brown", "843C0C"),
    ("blanco", "595959"),
    ("opa", "595959"),
    ("negro", "000000"),
    ("gris", "595959"),
    ("morado", "7030A0"),
    ("violeta", "7030A0"),
    ("rosa", "C04090"),
    ("fucsia", "FF00FF"),
    ("dorado", "BF8F00"),
    ("gold", "BF8F00"),
    ("beige", "A67C00"),
    ("crema", "A67C00"),
)


def logo_path() -> Optional[Path]:
    """PNG recomendado: Excel/openpyxl no incrusta WebP de forma nativa."""
    for rel in ("public/logo.png", "logo.png"):
        p = Path(resource_path(rel))
        if p.is_file():
            return p
    local = Path(__file__).resolve().parent.parent / "public" / "logo.png"
    return local if local.is_file() else None


def insertar_logo(ws, anchor: str = "A1", *, ancho_px: int = 130, alto_px: int = 48) -> bool:
    """Inserta logo superior izquierdo. Retorna True si se encontró el archivo."""
    path = logo_path()
    if path is None:
        return False
    img = XLImage(str(path))
    if img.width and img.height:
        escala = min(ancho_px / img.width, alto_px / img.height, 1.5)
        img.width = int(img.width * escala)
        img.height = int(img.height * escala)
    ws.add_image(img, anchor)
    fila = int(re.sub(r"[^0-9]", "", anchor) or "1")
    for r in range(fila, fila + 3):
        ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 16)
    return True


def color_hex_desde_nombre(nombre: str) -> str:
    """Infiera color de fuente a partir del maestro Color (ej. Azulino 660 → azul)."""
    n = (nombre or "").lower().strip()
    if not n:
        return "000000"
    for clave, hex_rgb in _COLOR_PALETTE:
        if clave in n:
            return hex_rgb
    return "000000"


def _font(color: str, *, bold: bool = False, size: int = 10) -> Font:
    return Font(color=color, bold=bold, size=size)


def estilo_celda_header(
    cell,
    *,
    advertencia: bool = False,
) -> None:
    cell.fill = FILL_HEADER_WARN if advertencia else FILL_HEADER
    cell.font = FONT_HEADER_WARN if advertencia else FONT_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def aplicar_fila_produccion(
    ws,
    row: int,
    reg: RegistroPesaje,
    idx: int,
    *,
    hora: str = "",
    verificacion: str = "ok Listo para imprimir",
) -> None:
    """
    Colorea la fila según el maestro Color (texto en tono del producto).
    Réplica amigable del Excel de planta: fondos por bloque de columnas.
    """
    tint = color_hex_desde_nombre(reg.color)
    f_prod = _font(tint)
    f_general = _font(tint)
    f_bruto = _font("C00000")
    f_neto = _font("000000", bold=True)
    f_idx = _font("C00000", bold=True)
    f_ok = _font("C00000", size=9)

    valores = (
        (1, idx, f_idx, None),
        (2, reg.nro_fardo, f_prod, FILL_PROD),
        (3, reg.cliente, f_prod, FILL_PROD),
        (4, reg.lote, f_prod, FILL_PROD),
        (5, reg.color, f_prod, FILL_PROD),
        (6, reg.denier, f_prod, FILL_PROD),
        (7, reg.corte, f_prod, FILL_PROD),
        (8, round(reg.peso_total, 2), f_general, FILL_PESO),
        (9, round(reg.tara_carreta, 2), f_general, FILL_PESO),
        (10, round(reg.tara_fardo, 2), f_general, FILL_PESO),
        (11, round(reg.peso_bruto, 2), f_bruto, FILL_PESO),
        (12, round(reg.peso_neto, 2), f_neto, FILL_PESO),
        (13, hora, f_general, FILL_HORA),
        (14, reg.operario, f_general, FILL_HORA),
        (15, reg.beteado or "1", f_general, None),
        (16, verificacion, f_ok, None),
    )

    for col, val, font, fill in valores:
        cell = ws.cell(row, col)
        if val is not None:
            cell.value = val
        cell.font = font
        if fill is not None:
            cell.fill = fill
        cell.border = BORDER
        if col == 1:
            cell.alignment = Alignment(horizontal="center")


def ajustar_columnas_dia(ws) -> None:
    widths = {
        "A": 5,
        "B": 8,
        "C": 18,
        "D": 14,
        "E": 18,
        "F": 6,
        "G": 9,
        "H": 10,
        "I": 11,
        "J": 10,
        "K": 10,
        "L": 10,
        "M": 8,
        "N": 14,
        "O": 9,
        "P": 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def estilo_fila_resumen(
    ws,
    row: int,
    *,
    produccion: bool,
    par: bool,
) -> None:
    fills = (FILL_KPI3, PatternFill()) if produccion else (PatternFill(), PatternFill())
    fill = fills[0] if par else fills[1]
    for col in range(2, 7):
        ws.cell(row, col).border = BORDER
        if produccion:
            ws.cell(row, col).fill = fill
