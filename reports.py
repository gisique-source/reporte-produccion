"""Exportación de reportes de producción a Excel (.xlsx) y PDF."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from db import format_fecha_editable
from models import RegistroPesaje

_HEADERS = (
    "Fecha/Hora",
    "Fardo",
    "Cliente",
    "Lote",
    "Color",
    "Dn",
    "Corte",
    "P.Total",
    "Tara Carr.",
    "Tara Fardo",
    "P.Bruto",
    "P.Neto",
    "Operario",
)


def _fila(reg: RegistroPesaje) -> tuple:
    return (
        reg.fecha_hora,
        reg.nro_fardo,
        reg.cliente,
        reg.lote,
        reg.color,
        reg.denier,
        reg.corte,
        round(reg.peso_total, 2),
        round(reg.tara_carreta, 2),
        round(reg.tara_fardo, 2),
        round(reg.peso_bruto, 2),
        round(reg.peso_neto, 2),
        reg.operario,
    )


def _totales(regs: Sequence[RegistroPesaje]) -> tuple[float, float, int]:
    bruto = sum(r.peso_bruto for r in regs)
    neto = sum(r.peso_neto for r in regs)
    return bruto, neto, len(regs)


def exportar_excel(
    regs: Sequence[RegistroPesaje],
    destino: str | Path,
    *,
    titulo: str,
    desde: date,
    hasta: date,
) -> Path:
    path = Path(destino)
    wb = Workbook()
    ws = wb.active
    ws.title = "Producción"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)

    ws["A1"] = titulo
    ws["A1"].font = title_font
    ws["A2"] = (
        f"Periodo: {format_fecha_editable(desde)} — {format_fecha_editable(hasta)}"
    )
    bruto, neto, cant = _totales(regs)
    ws["A3"] = f"Fardos: {cant}  |  Bruto: {bruto:,.2f} kg  |  Neto: {neto:,.2f} kg"

    start_row = 5
    for col, h in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, reg in enumerate(regs, start=1):
        for col, val in enumerate(_fila(reg), start=1):
            ws.cell(row=start_row + i, column=col, value=val)

    widths = (18, 8, 22, 14, 14, 8, 8, 10, 10, 10, 10, 10, 14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    return path


def exportar_pdf(
    regs: Sequence[RegistroPesaje],
    destino: str | Path,
    *,
    titulo: str,
    desde: date,
    hasta: date,
) -> Path:
    path = Path(destino)
    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(titulo, styles["Title"]),
        Paragraph(
            f"Periodo: {format_fecha_editable(desde)} — {format_fecha_editable(hasta)}",
            styles["Normal"],
        ),
    ]
    bruto, neto, cant = _totales(regs)
    story.append(
        Paragraph(
            f"Fardos: {cant} &nbsp;&nbsp; Bruto: {bruto:,.2f} kg &nbsp;&nbsp; "
            f"Neto: {neto:,.2f} kg",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6 * mm))

    data: list[list] = [list(_HEADERS)]
    for reg in regs:
        data.append([str(x) for x in _fila(reg)])

    if len(data) == 1:
        data.append(["(sin registros)"] + [""] * (len(_HEADERS) - 1))

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF3F8")]),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return path


def nombre_sugerido(prefijo: str, desde: date, hasta: date, ext: str) -> str:
    if desde == hasta:
        return f"{prefijo}_{desde.isoformat()}.{ext}"
    return f"{prefijo}_{desde.isoformat()}_{hasta.isoformat()}.{ext}"
