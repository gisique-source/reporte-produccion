"""Importación masiva desde libro Excel (.xlsx / .xlsm) tipo ETIQUETA EXTRUSORA."""

from __future__ import annotations

import calendar
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Optional

from bulk_import import FilaImport, parsear_matriz
from catalog import CatalogoMaestros
from config import TARA_CARRETA_KG, TARA_FARDO_KG

_DIA_SHEET = re.compile(r"^dia\s*0*(\d{1,2})$", re.I)
_MESES = {
    "ene": 1, "enero": 1, "jan": 1, "january": 1,
    "feb": 2, "febrero": 2, "february": 2,
    "mar": 3, "marzo": 3, "march": 3,
    "abr": 4, "abril": 4, "apr": 4, "april": 4,
    "may": 5, "mayo": 5,
    "jun": 6, "junio": 6, "june": 6,
    "jul": 7, "julio": 7, "july": 7,
    "ago": 8, "agosto": 8, "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "septiembre": 9, "september": 9,
    "oct": 10, "octubre": 10, "october": 10,
    "nov": 11, "noviembre": 11, "november": 11,
    "dic": 12, "diciembre": 12, "dec": 12, "december": 12,
}


@dataclass
class DiaExcel:
    """Una hoja Dia NN del libro mensual."""

    dia: int
    hoja: str
    fecha_en_hoja: Optional[date]
    filas: list[FilaImport] = field(default_factory=list)
    aviso: str = ""
    mapeo: dict[str, int] = field(default_factory=dict)

    @property
    def n_filas(self) -> int:
        return len(self.filas)

    @property
    def n_ok(self) -> int:
        return sum(1 for f in self.filas if f.lista_para_importar)

    @property
    def n_faltantes(self) -> int:
        return sum(1 for f in self.filas if f.tiene_faltantes)

    @property
    def n_errores(self) -> int:
        return sum(1 for f in self.filas if f.errores)


@dataclass
class LibroExcelImport:
    path: str
    anio_sugerido: int
    mes_sugerido: int
    dias: list[DiaExcel]
    alertas: list[str] = field(default_factory=list)
    origen_periodo: str = ""

    def dias_con_datos(self) -> list[DiaExcel]:
        return [d for d in self.dias if d.n_filas > 0]


def celda_a_texto(value: Any) -> str:
    """Normaliza celdas openpyxl (fecha, hora, float, #N/A) a str para el parser."""
    if value is None:
        return ""
    if isinstance(value, float) and value != value:  # NaN
        return ""
    if isinstance(value, str):
        s = value.strip()
        if s.upper() in ("#N/A", "#¡N/A!", "#VALUE!", "#REF!", "#DIV/0!"):
            return ""
        return s
    if isinstance(value, datetime):
        # Excel a veces guarda hora >24h como datetime 1900-01-01 …
        if value.year <= 1900:
            return f"{value.hour}:{value.minute:02d}"
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return f"{value.hour}:{value.minute:02d}"
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value).strip()


def _inferir_periodo_nombre(path: str) -> Optional[tuple[int, int]]:
    """
    Ej.: 'ETIQUETA EXTRUSORA 05-26- MAY.xlsm' → (2026, 5)
         'ETIQUETA 05-2026.xlsx' → (2026, 5)
    """
    stem = Path(path).stem
    # MM-YY o MM-YYYY
    m = re.search(r"(?<!\d)(0?[1-9]|1[0-2])[-_/ ](\d{2}|\d{4})(?!\d)", stem)
    mes_txt = None
    anio = None
    mes = None
    if m:
        mes = int(m.group(1))
        yy = int(m.group(2))
        anio = yy if yy >= 100 else 2000 + yy
    low = stem.casefold()
    for nombre, num in _MESES.items():
        if re.search(rf"(?<![a-záéíóú]){re.escape(nombre)}(?![a-záéíóú])", low):
            mes_txt = num
            break
    if mes is None and mes_txt:
        mes = mes_txt
    elif mes is not None and mes_txt and mes != mes_txt:
        # Preferir el número explícito si hay conflicto
        pass
    if mes is None or anio is None:
        return None
    return anio, mes


def _periodo_desde_fechas(fechas: list[date]) -> Optional[tuple[int, int, str]]:
    if not fechas:
        return None
    pairs = [(f.year, f.month) for f in fechas]
    (anio, mes), _n = Counter(pairs).most_common(1)[0]
    return anio, mes, "fechas en hojas Dia"


def validar_dia_en_mes(anio: int, mes: int, dia: int) -> Optional[str]:
    """None si es válido; mensaje de alerta si no existe (p. ej. 31 en febrero)."""
    if mes < 1 or mes > 12:
        return f"Mes {mes} inválido."
    max_d = calendar.monthrange(anio, mes)[1]
    if dia < 1 or dia > max_d:
        from db import nombre_mes

        return (
            f"El día {dia:02d} no existe en "
            f"{nombre_mes(mes)} {anio} "
            f"(ese mes tiene {max_d} días)."
        )
    return None


def leer_libro_excel(
    path: str,
    catalogo: CatalogoMaestros,
    *,
    tara_c_default: float = TARA_CARRETA_KG,
    tara_f_default: float = TARA_FARDO_KG,
) -> LibroExcelImport:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Falta openpyxl en el entorno.") from exc

    wb = load_workbook(path, data_only=True, keep_vba=False, read_only=True)
    try:
        dias: list[DiaExcel] = []
        fechas_hoja: list[date] = []

        for name in wb.sheetnames:
            m = _DIA_SHEET.match(name.strip())
            if not m:
                continue
            dia_n = int(m.group(1))
            if dia_n < 1 or dia_n > 31:
                continue
            ws = wb[name]
            fecha_hoja = _leer_fecha_produccion(ws)
            if fecha_hoja:
                fechas_hoja.append(fecha_hoja)

            matriz = _matriz_desde_hoja(ws)
            if not matriz:
                dias.append(
                    DiaExcel(
                        dia=dia_n,
                        hoja=name,
                        fecha_en_hoja=fecha_hoja,
                        aviso="Sin tabla de producción",
                    )
                )
                continue

            resultado = parsear_matriz(
                matriz,
                catalogo,
                tara_c_default=tara_c_default,
                tara_f_default=tara_f_default,
            )
            dias.append(
                DiaExcel(
                    dia=dia_n,
                    hoja=name,
                    fecha_en_hoja=fecha_hoja,
                    filas=list(resultado.filas),
                    aviso=resultado.aviso,
                    mapeo=dict(resultado.mapeo),
                )
            )

        dias.sort(key=lambda d: d.dia)
        alertas: list[str] = []
        origen = ""

        periodo_nom = _inferir_periodo_nombre(path)
        periodo_fec = _periodo_desde_fechas(fechas_hoja)

        if periodo_fec:
            anio, mes, origen = periodo_fec
        elif periodo_nom:
            anio, mes = periodo_nom
            origen = "nombre de archivo"
        else:
            hoy = date.today()
            anio, mes = hoy.year, hoy.month
            origen = "fecha actual (no se detectó mes/año)"
            alertas.append(
                "No se pudo detectar mes/año del archivo. "
                "Seleccione el período manualmente antes de importar."
            )

        if periodo_nom and periodo_fec:
            an_nom, mes_nom = periodo_nom
            if (an_nom, mes_nom) != (periodo_fec[0], periodo_fec[1]):
                alertas.append(
                    f"Disparidad: el nombre sugiere {mes_nom:02d}/{an_nom} "
                    f"pero las hojas indican mayoritariamente "
                    f"{periodo_fec[1]:02d}/{periodo_fec[0]}. "
                    f"Se usa {mes:02d}/{anio} ({origen})."
                )

        # Fechas de hoja con año raro respecto al período
        for d in dias:
            if d.fecha_en_hoja and (
                d.fecha_en_hoja.year != anio or d.fecha_en_hoja.month != mes
            ):
                if d.n_filas > 0:
                    alertas.append(
                        f"{d.hoja}: fecha en celda = "
                        f"{d.fecha_en_hoja.isoformat()} "
                        f"(distinta del período {mes:02d}/{anio})."
                    )

        # Días del libro que no existen en el mes sugerido
        for d in dias:
            if d.n_filas == 0:
                continue
            msg = validar_dia_en_mes(anio, mes, d.dia)
            if msg:
                alertas.append(f"{d.hoja}: {msg}")

        # Dedup alertas preservando orden
        seen: set[str] = set()
        alertas_u: list[str] = []
        for a in alertas:
            if a not in seen:
                seen.add(a)
                alertas_u.append(a)

        return LibroExcelImport(
            path=path,
            anio_sugerido=anio,
            mes_sugerido=mes,
            dias=dias,
            alertas=alertas_u,
            origen_periodo=origen,
        )
    finally:
        wb.close()


def _leer_fecha_produccion(ws: Any) -> Optional[date]:
    """Busca 'Fecha de Produccion' en las primeras filas o celda típica M5."""
    try:
        for row in ws.iter_rows(min_row=1, max_row=8, max_col=20, values_only=True):
            for i, cell in enumerate(row):
                if isinstance(cell, str) and re.search(
                    r"fecha\s*(de\s*)?produc", cell, re.I
                ):
                    for j in range(i + 1, min(len(row), i + 5)):
                        d = _as_date(row[j])
                        if d:
                            return d
            # Columna M (índice 12) suele traer la fecha en plantillas Gexim
            if len(row) > 12:
                d = _as_date(row[12])
                if d and d.year >= 2000:
                    return d
    except Exception:  # noqa: BLE001
        return None
    return None


def _as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _matriz_desde_hoja(ws: Any) -> list[list[str]]:
    """
    Localiza la fila de encabezado (Fardo/Cliente/…) y convierte filas a str.
    Omite la columna A si es solo índice correlativo vacío de encabezado.
    """
    raw_rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=250, max_col=20, values_only=True):
        raw_rows.append(list(row))

    header_idx = None
    for i, row in enumerate(raw_rows):
        texts = [celda_a_texto(c) for c in row]
        joined = " ".join(texts).casefold()
        if "fardo" in joined and "cliente" in joined and "lote" in joined:
            header_idx = i
            break
    if header_idx is None:
        return []

    # Recortar columnas vacías a la izquierda si el encabezado empieza en col B
    header = [celda_a_texto(c) for c in raw_rows[header_idx]]
    start_col = 0
    if not header[0] and any(header[1:]):
        # Columna A vacía en encabezado → datos empiezan en B, pero el índice
        # numérico de Excel a veces está en A; dejamos todas las columnas
        # para que el mapper encuentre "Fardo" en la posición correcta.
        start_col = 0

    matriz: list[list[str]] = []
    for row in raw_rows[header_idx:]:
        cells = [celda_a_texto(c) for c in row[start_col:]]
        if any(cells):
            matriz.append(cells)
    return matriz


def fechas_efectivas(
    libro: LibroExcelImport, anio: int, mes: int
) -> dict[int, tuple[Optional[date], Optional[str]]]:
    """
    Para cada día del libro: (fecha_ok | None, alerta | None).
    """
    out: dict[int, tuple[Optional[date], Optional[str]]] = {}
    for d in libro.dias:
        msg = validar_dia_en_mes(anio, mes, d.dia)
        if msg:
            out[d.dia] = (None, msg)
        else:
            out[d.dia] = (date(anio, mes, d.dia), None)
    return out
